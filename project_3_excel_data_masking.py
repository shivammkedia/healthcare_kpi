"""
Excel Data Masking and Anonymization Utility

This utility provides enterprise-grade data masking for sensitive Excel files, particularly
useful for healthcare and financial datasets. It:
1. Reads Excel files and analyzes data types
2. Generates realistic but fake data using Faker
3. Preserves data structure and types while anonymizing values
4. Applies intelligent masking based on column names and data types
5. Generates detailed reports and formatted output files

Technologies: Pandas, Openpyxl, Faker, Numpy
"""

import pandas as pd
import numpy as np
from pathlib import Path
from faker import Faker
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import uuid
import warnings

warnings.filterwarnings('ignore')


class ExcelMasker:
    """
    A utility class to read Excel files, analyze data types, 
    and create masked/anonymized versions of the data.
    
    This class is designed for healthcare and sensitive data scenarios where
    data structure must be preserved but actual values need to be anonymized.
    
    Attributes:
        input_file (Path): Path to the input Excel file
        fake (Faker): Faker instance for generating fake data
        df (DataFrame): Original dataframe loaded from Excel
        masked_df (DataFrame): Anonymized version of the dataframe
        data_types (dict): Dictionary of column data type information
    """

    def __init__(self, input_file_path):
        """
        Initialize ExcelMasker with input file path.
        
        Args:
            input_file_path (str): Path to the Excel file to be masked
            
        Raises:
            FileNotFoundError: If the specified file does not exist
        """
        self.input_file = Path(input_file_path)
        self.fake = Faker()
        self.df = None
        self.data_types = {}

        if not self.input_file.exists():
            raise FileNotFoundError(f"File not found: {input_file_path}")

    def read_excel(self):
        """
        Read the Excel file and return DataFrame.
        
        Returns:
            DataFrame: The loaded dataframe, or None if error occurred
        """
        try:
            self.df = pd.read_excel(self.input_file)
            print(f"✓ Successfully read Excel file: {self.input_file.name}")
            print(f"  - Rows: {len(self.df)}")
            print(f"  - Columns: {len(self.df.columns)}\n")
            return self.df
        except Exception as e:
            print(f"✗ Error reading Excel file: {e}")
            return None

    def analyze_data_types(self):
        """
        Analyze and display data types of each column.
        
        Examines each column for:
        - Data type (int, float, object, datetime, bool)
        - Non-null and null counts
        - Sample values
        
        Returns:
            dict: Dictionary mapping column names to their type information
        """
        if self.df is None:
            print("✗ Please read the Excel file first using read_excel()")
            return None

        print("=" * 80)
        print("DATA TYPE ANALYSIS")
        print("=" * 80)

        for column in self.df.columns:
            dtype = self.df[column].dtype
            sample_value = self.df[column].dropna().iloc[0] if len(self.df[column].dropna()) > 0 else "N/A"
            non_null_count = self.df[column].notna().sum()
            null_count = self.df[column].isna().sum()

            self.data_types[column] = {
                'dtype': dtype,
                'non_null': non_null_count,
                'null': null_count,
                'sample': sample_value
            }

            print(f"\n{column}:")
            print(f"  Data Type: {dtype}")
            print(f"  Non-null: {non_null_count} | Null: {null_count}")
            print(f"  Sample Value: {sample_value}")

        print("\n" + "=" * 80 + "\n")
        return self.data_types

    def generate_masked_data(self):
        """
        Generate masked/anonymized data based on data types and column names.
        
        Uses intelligent pattern matching on column names to determine masking strategy:
        - ID columns → UUID-based identifiers
        - Email columns → Faker emails
        - Phone columns → Faker phone numbers
        - Name columns → Faker names
        - Address columns → Faker addresses
        - Date columns → Random dates in range
        - Amount/Price columns → Random floats
        - Integer columns → Random integers
        - Boolean columns → Random true/false
        - Other columns → Generic masked values
        
        Returns:
            DataFrame: The masked/anonymized dataframe
        """
        if self.df is None:
            print("✗ Please read the Excel file first using read_excel()")
            return None

        print("Generating masked data...")
        masked_df = pd.DataFrame()

        for column in self.df.columns:
            dtype = self.df[column].dtype
            column_size = len(self.df)

            # Generate masked data based on column data type and name
            if 'id' in column.lower() or 'code' in column.lower():
                # Generate fake IDs/codes using UUID
                masked_df[column] = [str(uuid.uuid4())[:8].upper() for _ in range(column_size)]

            elif 'email' in column.lower():
                # Generate fake emails
                masked_df[column] = [self.fake.email() for _ in range(column_size)]

            elif 'phone' in column.lower() or 'mobile' in column.lower():
                # Generate fake phone numbers
                masked_df[column] = [self.fake.phone_number() for _ in range(column_size)]

            elif 'name' in column.lower():
                # Generate fake names
                masked_df[column] = [self.fake.name() for _ in range(column_size)]

            elif 'address' in column.lower() or 'location' in column.lower():
                # Generate fake addresses
                masked_df[column] = [self.fake.address() for _ in range(column_size)]

            elif 'date' in column.lower() or pd.api.types.is_datetime64_any_dtype(dtype):
                # Generate random dates within a reasonable range
                start_date = datetime(2020, 1, 1)
                end_date = datetime(2024, 12, 31)
                masked_df[column] = [
                    start_date + timedelta(days=int(np.random.random() * (end_date - start_date).days))
                    for _ in range(column_size)
                ]

            elif 'amount' in column.lower() or 'price' in column.lower() or 'cost' in column.lower() or \
                    pd.api.types.is_float_dtype(dtype):
                # Generate random amounts/prices in realistic range
                masked_df[column] = np.random.uniform(100, 10000, column_size)

            elif pd.api.types.is_integer_dtype(dtype):
                # Generate random integers
                masked_df[column] = np.random.randint(1, 1000, column_size)

            elif pd.api.types.is_bool_dtype(dtype):
                # Generate random boolean values
                masked_df[column] = np.random.choice([True, False], column_size)

            else:
                # For other types, generate generic masked values
                masked_df[column] = [f"MASKED_{i}" for i in range(column_size)]

        self.masked_df = masked_df
        print(f"✓ Masked data generated with {len(masked_df)} rows and {len(masked_df.columns)} columns\n")
        return masked_df

    def save_masked_excel(self, output_file_path=None):
        """
        Save the masked data to a new Excel file.
        
        Args:
            output_file_path (str, optional): Path for output file. 
                                             Defaults to {input_name}_masked.xlsx
        
        Returns:
            bool: True if successful, False otherwise
        """
        if not hasattr(self, 'masked_df') or self.masked_df is None:
            print("✗ Please generate masked data first using generate_masked_data()")
            return False

        if output_file_path is None:
            output_file_path = self.input_file.parent / f"{self.input_file.stem}_masked.xlsx"

        try:
            # Save to Excel
            self.masked_df.to_excel(output_file_path, index=False)

            # Format the Excel file with professional styling
            self._format_excel(output_file_path)

            print(f"✓ Masked Excel file saved: {output_file_path}")
            print(f"  File size: {Path(output_file_path).stat().st_size / 1024:.2f} KB")
            return True
        except Exception as e:
            print(f"✗ Error saving masked Excel file: {e}")
            return False

    def _format_excel(self, file_path):
        """
        Format the Excel file with colors and styling.
        
        Applies:
        - Blue header row with white bold text
        - Auto-adjusted column widths
        - Centered alignment
        
        Args:
            file_path (str): Path to the Excel file to format
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Style the header row with professional blue color
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Adjust column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
        except Exception as e:
            print(f"Warning: Could not format Excel file: {e}")

    def display_comparison(self):
        """
        Display side-by-side comparison of original and masked data.
        
        Shows the first 3 rows to demonstrate how data was transformed
        while preserving structure and data types.
        """
        if self.df is None or not hasattr(self, 'masked_df'):
            print("✗ Please read the Excel file and generate masked data first")
            return

        print("=" * 80)
        print("ORIGINAL vs MASKED DATA COMPARISON (First 3 rows)")
        print("=" * 80 + "\n")

        for idx in range(min(3, len(self.df))):
            print(f"Row {idx + 1}:")
            print("-" * 80)
            for column in self.df.columns:
                original = self.df.loc[idx, column]
                masked = self.masked_df.loc[idx, column]
                print(f"  {column}:")
                print(f"    Original: {original}")
                print(f"    Masked:   {masked}")
            print()

    def generate_report(self, output_file_path=None):
        """
        Generate a detailed report about the masking process.
        
        Creates a text report containing:
        - File paths and timestamps
        - Data type analysis for each column
        - Masking strategies used
        
        Args:
            output_file_path (str, optional): Path for report file.
                                             Defaults to {input_name}_masking_report.txt
        
        Returns:
            bool: True if successful, False otherwise
        """
        if output_file_path is None:
            output_file_path = self.input_file.parent / f"{self.input_file.stem}_masking_report.txt"

        try:
            with open(output_file_path, 'w') as f:
                f.write("=" * 80 + "\n")
                f.write("EXCEL DATA MASKING REPORT\n")
                f.write("=" * 80 + "\n\n")

                f.write(f"Input File: {self.input_file}\n")
                f.write(f"Output File: {self.input_file.parent / f'{self.input_file.stem}_masked.xlsx'}\n")
                f.write(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

                f.write("DATA TYPE ANALYSIS:\n")
                f.write("-" * 80 + "\n")
                for column, info in self.data_types.items():
                    f.write(f"\n{column}:\n")
                    f.write(f"  Data Type: {info['dtype']}\n")
                    f.write(f"  Non-null values: {info['non_null']}\n")
                    f.write(f"  Null values: {info['null']}\n")
                    f.write(f"  Sample value: {info['sample']}\n")

                f.write("\n" + "=" * 80 + "\n")

            print(f"✓ Report saved: {output_file_path}\n")
            return True
        except Exception as e:
            print(f"✗ Error generating report: {e}")
            return False


# Example usage
def main():
    """
    Main function demonstrating how to use the ExcelMasker class.
    
    Workflow:
    1. Initialize masker with Excel file
    2. Read and analyze the file
    3. Generate masked/anonymized data
    4. Display comparison
    5. Save masked Excel file
    6. Generate masking report
    """

    # Define file paths
    input_file = "E:/python_files/healthcare_dataset.xlsx"

    try:
        # Create masker instance
        masker = ExcelMasker(input_file)

        # Step 1: Read the Excel file
        print("\n" + "=" * 80)
        print("STEP 1: READING EXCEL FILE")
        print("=" * 80)
        masker.read_excel()

        # Step 2: Analyze data types
        print("STEP 2: ANALYZING DATA TYPES")
        print("=" * 80)
        masker.analyze_data_types()

        # Step 3: Generate masked data
        print("\nSTEP 3: GENERATING MASKED DATA")
        print("=" * 80)
        masker.generate_masked_data()

        # Step 4: Display comparison
        print("\nSTEP 4: DISPLAYING COMPARISON")
        print("=" * 80)
        masker.display_comparison()

        # Step 5: Save masked Excel file
        print("\nSTEP 5: SAVING MASKED EXCEL FILE")
        print("=" * 80)
        masker.save_masked_excel()

        # Step 6: Generate report
        print("\nSTEP 6: GENERATING REPORT")
        print("=" * 80)
        masker.generate_report()

        print("\n" + "=" * 80)
        print("MASKING PROCESS COMPLETED SUCCESSFULLY!")
        print("=" * 80 + "\n")

    except Exception as e:
        print(f"\n✗ Error: {e}\n")


if __name__ == "__main__":
    main()
